"""
Catalogue data files already on disk, classifying them from their contents.

reclassify_manifest.py labels files from the link text a website gave them, which
fails when the filename is meaningless (Minnesota's download handler names 115 of
its files 000574.xlsx) and needs a manifest to exist at all. This walks a
directory instead, builds a content signature per file (sheets, header columns,
sample values, row count, years) and classifies from that.

Columns are split by how far to trust them:
    llm_category/llm_topic/llm_confidence   the model reading the signature
    entity_levels/verified_dims/dims_source deterministic, from the columns

Row count matters: without it the model called a 33,706-row extract a "format
template" because its column names read like a spec. Anything above
TEMPLATE_MAX_ROWS is forced out of that category.

    python scripts/common/catalog_local.py data/raw/mn --out data/catalog/mn.csv
    python scripts/common/catalog_local.py data/raw/mo --out data/catalog/mo.csv --no-llm
"""

import argparse
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import tqdm

from common import llm_assist as L
from common.verify_dims import inspect as inspect_dims

DATA_EXTS = {".xlsx", ".xlsm", ".xls", ".csv", ".txt", ".tab"}
MAX_SIG_BYTES = 250_000_000
TEMPLATE_MAX_ROWS = 50   # above this a file is data, whatever its column names look like
VALID_CATEGORIES = {"assessment", "enrollment", "attendance", "graduation", "discipline",
                    "staff", "finance", "directory", "template", "other"}
YEAR_RE = re.compile(r"(19|20)\d{2}")

# Entity level is a plain column-name question, so decide it deterministically
# rather than asking the model (which left it blank on files whose headers plainly
# said "District" and "School").
LEVEL_PATTERNS = {
    # match a bare "School"/"District" column too, but never "School Year"/"School Type"
    "school": re.compile(r"\b(school|building|site)([_ ]?(name|number|code|id))?\b", re.I),
    "district": re.compile(r"\b(district|corporation|lea|charter)([_ ]?(name|number|code|id))?\b", re.I),
    "county": re.compile(r"\bcounty([_ ]?(name|number|code))?\b", re.I),
    "state": re.compile(r"\b(statewide|state[_ ]?(name|total|level))\b", re.I),
}
NOT_A_LEVEL = re.compile(r"\bschool[_ ]?(year|type|class|day|age|district)\b", re.I)


def _levels_from_columns(columns, sample):
    found = set()
    # drop columns that merely contain the word (e.g. "School Year") before matching
    cols = [c for c in list(columns)[:40] if not NOT_A_LEVEL.fullmatch(str(c).strip())]
    hay = " | ".join(cols)
    for lvl, pat in LEVEL_PATTERNS.items():
        if pat.search(hay):
            found.add(lvl)
    # a "Statewide" row value also implies state coverage
    if any("statewide" in str(v).lower() for v in list(sample)[:40]):
        found.add("state")
    return sorted(found)

SIG_PROMPT = """You are cataloguing public K-12 education data files for a research project.
Each entry below is a real file: its filename, its spreadsheet sheet names, and its
actual column headers (plus sample values). Classify from the CONTENT, not the filename —
many filenames are meaningless codes.

For each numbered file return:
  "n"        the file's number
  "category" exactly one of: {cats}
  "topic"    a short human label for what the file contains, max 8 words
             (e.g. "MCA math proficiency by student group", "district per-pupil expenditures")
  "levels"   which entity levels the columns imply, any of: state, county, district, school
             (look for District Name/Number, School Name/Number, County columns)
  "conf"     your confidence 0-1 that this classification is right

Category guidance:
- assessment = test/exam results (MCA, MTAS, ILEARN, IREAD, MAP, SAT, ACT, WIDA, ACCESS, EOC, proficiency, scale score)
- enrollment = student counts, membership, demographics, free/reduced-price meal counts
- attendance = attendance rate, absenteeism, ADA/ADM, mobility
- staff = teachers, faculty, educators, certification, salary, personnel, staff ratios
- finance = expenditures, revenue, per-pupil spending, budgets, funding, levy, valuation
- graduation = graduation/dropout/cohort/completion/college-going
- discipline = suspensions, expulsions, incidents, referrals, safety
- directory = school/district lists, codes, addresses, calendars
- template = a BLANK submission template or file-layout spec. Only use this when
  "data rows" is very small (roughly < 20). A file with thousands of data rows is real
  data even if its column names read like a specification.
- other = anything else

Return ONLY JSON: {{"results":[ ... ]}}

FILES:
{files}"""


def _signature(path):
    """Return (sheets, columns, sample_values, years, n_rows) read cheaply.

    n_rows matters: it is what separates a real dataset from a blank submission
    template. Without it the model called Minnesota's 33,705-row SOLOM extract a
    "format template" purely because its column names read like a spec.
    """
    p = Path(path)
    sheets, columns, sample, n_rows = [], [], [], None
    try:
        if p.stat().st_size > MAX_SIG_BYTES:
            return sheets, columns, sample, [], n_rows
        suf = p.suffix.lower()
        if suf in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            try:
                sheets = list(wb.sheetnames)[:8]
                # prefer a sheet that looks like data, not notes/definitions
                order = sorted(wb.sheetnames,
                               key=lambda s: any(k in s.lower() for k in
                                                 ("definition", "note", "info", "overview", "readme")))
                for sn in order[:3]:
                    ws = wb[sn]
                    for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True)):
                        cells = [str(c).strip() for c in (row or []) if c is not None]
                        if len(cells) > 3 and not columns:
                            columns = cells[:40]
                        elif columns and cells:
                            sample.extend(cells[:12])
                        if i > 12 or len(sample) > 60:
                            break
                    if columns:
                        n_rows = ws.max_row
                        break
                if n_rows is None and wb.sheetnames:
                    # no header row found (near-empty / notes-only file) — still report
                    # the row count so a template stays distinguishable from data
                    n_rows = wb[order[0]].max_row
            finally:
                wb.close()
        elif suf == ".xls":
            # openpyxl cannot read the legacy BIFF format (386 such files here);
            # without this they arrived at the model with no columns at all.
            import xlrd
            bk = xlrd.open_workbook(p, on_demand=True)
            try:
                sheets = bk.sheet_names()[:8]
                for sn in sheets[:3]:
                    sh = bk.sheet_by_name(sn)
                    for i in range(min(sh.nrows, 12)):
                        cells = [str(c).strip() for c in sh.row_values(i) if str(c).strip()]
                        if len(cells) > 3 and not columns:
                            columns = cells[:40]
                        elif columns and cells:
                            sample.extend(cells[:12])
                        if len(sample) > 60:
                            break
                    if columns:
                        n_rows = sh.nrows
                        break
            finally:
                bk.release_resources()
        elif suf in (".csv", ".txt", ".tab"):
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                head = f.read(32768)
                delim = "\t" if head.count("\t") > head.count(",") else ","
                lines = head.splitlines()
                if lines:
                    columns = [c.strip() for c in lines[0].split(delim)][:40]
                    for ln in lines[1:6]:
                        sample.extend([c.strip() for c in ln.split(delim)][:12])
                # count remaining rows without loading the file into memory
                n_rows = len(lines) + sum(1 for _ in f)
    except Exception:
        pass
    blob = " ".join([p.name] + sheets + columns + sample[:40])
    years = sorted({m.group(0) for m in YEAR_RE.finditer(blob)})[:6]
    return sheets, columns, sample, years, n_rows


def _entry_text(i, rec):
    parts = [f"{i}. file: {rec['name']}"]
    if rec.get("n_rows") is not None:
        parts.append(f"   data rows: {rec['n_rows']}")
    if rec["sheets"]:
        parts.append(f"   sheets: {', '.join(rec['sheets'][:6])}")
    if rec["columns"]:
        parts.append(f"   columns: {', '.join(rec['columns'][:22])}")
    if rec["sample"]:
        parts.append(f"   values: {', '.join(rec['sample'][:12])}")
    return "\n".join(p[:400] for p in parts)


def classify(records, batch=12, verbose=True):
    """records: list of dicts with name/sheets/columns/sample. Returns aligned results."""
    out = [None] * len(records)
    for start in range(0, len(records), batch):
        chunk = records[start:start + batch]
        prompt = SIG_PROMPT.format(
            cats="assessment, enrollment, attendance, graduation, discipline, staff, "
                 "finance, directory, template, other",
            files="\n".join(_entry_text(i + 1, r) for i, r in enumerate(chunk)))
        try:
            data = L._parse_json(L._generate(prompt, num_predict=1800))
        except Exception as e:
            if verbose:
                tqdm.tqdm.write(f"    llm batch @{start} failed: {str(e)[:70]}")
            continue
        results = (data or {}).get("results") if isinstance(data, dict) else data
        if not isinstance(results, list):
            continue
        for rec in results:
            if not isinstance(rec, dict):
                continue
            try:
                n = int(rec.get("n", 0)) - 1
            except (TypeError, ValueError):
                continue
            if not (0 <= n < len(chunk)):
                continue
            lv = rec.get("levels") or []
            if isinstance(lv, str):
                lv = [lv]
            cat = str(rec.get("category", "other")).lower().strip()
            # constrain to the known set — the model occasionally invents a spelling
            # ("graduration") which would otherwise become its own category
            if cat not in VALID_CATEGORIES:
                cat = "other"
            # Guard rail: the model reaches for "template" whenever column names read
            # like a spec, even on a 33,000-row extract. Row count settles it.
            nr = chunk[n].get("n_rows")
            if cat == "template" and nr is not None and nr > TEMPLATE_MAX_ROWS:
                cat = "other"
            out[start + n] = {
                "category": cat,
                "topic": str(rec.get("topic", ""))[:90],
                "levels": "|".join(str(x).lower().strip() for x in lv
                                   if str(x).lower().strip() in
                                   ("state", "county", "district", "school")),
                "conf": rec.get("conf", ""),
            }
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("root")
    ap.add_argument("--out", required=True)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--no-llm", action="store_true")
    ap.add_argument("--min-kb", type=int, default=1)
    args = ap.parse_args()

    root = Path(args.root)
    files = [p for p in sorted(root.rglob("*"))
             if p.is_file() and p.suffix.lower() in DATA_EXTS
             and p.stat().st_size >= args.min_kb * 1024]
    if args.limit:
        files = files[:args.limit]
    print(f"{len(files)} data files under {root}")

    records = []
    for p in tqdm.tqdm(files, desc="reading signatures"):
        sheets, columns, sample, years, n_rows = _signature(p)
        dims, dsrc = inspect_dims(p)
        records.append({
            "path": str(p), "name": p.name,
            "size_mb": round(p.stat().st_size / 1e6, 2),
            "sheets": sheets, "columns": columns, "sample": sample, "years": years,
            "n_rows": n_rows,
            "dims": sorted(dims), "dims_source": dsrc,
            "levels": _levels_from_columns(columns, sample),
        })

    results = [None] * len(records)
    if not args.no_llm:
        if not L.available():
            print(f"Ollama/{L.MODEL} unavailable — writing signatures only "
                  f"(re-run without --no-llm once `ollama serve` is up).")
        else:
            print(f"classifying with {L.MODEL} ...")
            results = classify(records)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["path", "size_mb", "sheets", "n_cols", "n_rows", "columns", "years",
                    "llm_category", "llm_topic", "llm_confidence",
                    "entity_levels", "verified_dims", "dims_source"])
        for rec, res in zip(records, results):
            w.writerow([
                rec["path"], rec["size_mb"], "|".join(rec["sheets"][:6]),
                len(rec["columns"]), rec["n_rows"] if rec["n_rows"] is not None else "",
                "|".join(rec["columns"][:25]), "|".join(rec["years"]),
                (res or {}).get("category", ""), (res or {}).get("topic", ""),
                (res or {}).get("conf", ""),
                "|".join(rec["levels"]), "|".join(rec["dims"]), rec["dims_source"],
            ])

    from collections import Counter
    print(f"\nwrote {args.out}")
    if any(results):
        print("categories:", dict(Counter((r or {}).get("category", "?") for r in results).most_common()))
    withdims = [r for r in records if r["dims"]]
    print(f"files with verified demographic columns: {len(withdims)}/{len(records)}")
    print("dimensions:", dict(Counter(d for r in withdims for d in r["dims"]).most_common()))


if __name__ == "__main__":
    main()
