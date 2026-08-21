"""
Normalise the Kansas annual assessment workbooks into one tidy CSV.

The eleven full files downloaded by ks/assessment.py are not a consistent series -
KSDE changed the layout roughly every other year:

    Each workbook is a TWO-YEAR window: 2023_2024 holds sheets `2023` and `2024`,
    2024_2025 holds `2024` and `2025`. Every year except the endpoints is therefore
    published twice, and the two copies are not always identical - KSDE revises. The
    later edition wins; rows the earlier edition has and the later one does not are
    kept (this is how the 2016-17 "Accountability" population survives).

    2014-15   placeholder, "No assessment data available" (assessment transition)
    2015-16   sheet AssessmentResults; program_year / Org_No / GradeName / GroupName
    2016-17   orglevel-first layout, adds bldgno / orgno, keeps Population
    2017-18   same minus Population
    2018-19   headers renamed to "Pct. Level 1", "Group", "Bldg. No."
    2019-20   renamed again: "Organization Level", "Pct. Level One", "GroupName"
    2020-21   placeholder, no data (COVID)
    2021-22   drops Pct Not Valid; Grade becomes "3rd Grade" rather than "3"
    2022-25   current layout: School Year / Org. No. / Bldg. No. / Student Subgroup

Two further wrinkles worth knowing before analysing:

  * 2015-16 and 2016-17 carry a `Population` column with both "Accountability" and
    "Report Card" rows - two different populations for the same cell, which is why
    those files have ~500k rows against ~300k elsewhere. Later years publish the
    report-card population only, so `population` is blank there rather than assumed.
  * grade labels are normalised to bare numbers plus "All Grades" / "HS".
  * gender is published for 2014-15 and 2017-18 only ("Males"/"Females"); the other
    eight years carry race, poverty, disability, EL and mobility but no gender.

Output: data/raw/ks/assessment/assessment_all_years.csv

    python scripts/ks/normalize_assessment.py
"""

import csv
import glob
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import tqdm

FULL_DIR = Path("data/raw/ks/assessment/full_files")
DEST = Path("data/raw/ks/assessment/assessment_all_years.csv")

CANON = ["school_year", "org_no", "bldg_no", "organization", "building",
         "student_group", "grade", "subject", "population",
         "pct_level_1", "pct_level_2", "pct_level_3", "pct_level_4", "pct_not_valid"]

# header (lowercased, punctuation stripped) -> canonical name
ALIASES = {
    "programyear": "school_year", "schoolyear": "school_year",
    "orgno": "org_no",
    "buildingnumber": "bldg_no", "bldgno": "bldg_no",
    "organization": "organization", "orglevel": "organization",
    "organizationlevel": "organization", "organizationbuilding": "organization",
    "building": "building",
    "groupname": "student_group", "group": "student_group",
    "studentsubgroup": "student_group",
    "gradename": "grade", "grade": "grade",
    "subject": "subject",
    "population": "population",
    "pclevelone": "pct_level_1", "pctlevel1": "pct_level_1", "pctlevelone": "pct_level_1",
    "pcleveltwo": "pct_level_2", "pctlevel2": "pct_level_2", "pctleveltwo": "pct_level_2",
    "pclevelthree": "pct_level_3", "pctlevel3": "pct_level_3", "pctlevelthree": "pct_level_3",
    "pclevelfour": "pct_level_4", "pctlevel4": "pct_level_4", "pctlevelfour": "pct_level_4",
    "pcnotvalid": "pct_not_valid", "pctnotvalid": "pct_not_valid",
    "pctnottested": "pct_not_valid", "pcnottested": "pct_not_valid",
}

GRADE_RE = re.compile(r"^\s*(\d+)")


def _key(h):
    return re.sub(r"[^a-z0-9]", "", str(h).lower())


def _grade(v):
    """'3', '3rd Grade', 'ALL', '13' -> '3', 'All Grades'.

    KSDE spells the all-grades aggregate three ways across the series: "All Grades",
    "ALL" (2015-16) and the bare code "13" (2017-19). 13 is not a real grade -
    Kansas tests 3-8, 10 and 11 - and in those files it is the most frequent value,
    which is what an aggregate row looks like.
    """
    s = str(v or "").strip()
    if not s:
        return ""
    low = s.lower()
    if "all" in low or s == "13":
        return "All Grades"
    if low in ("hs", "high school") or "high" in low:
        return "HS"
    m = GRADE_RE.match(s)
    return m.group(1) if m else s


def read_workbook(path):
    """Yield canonical dict rows from one workbook, whatever its layout."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            try:
                header = next(it)
            except StopIteration:
                continue
            cols = [_key(h) for h in header]
            # a placeholder sheet ("No assessment data available") has no real header
            mapping = {i: ALIASES[c] for i, c in enumerate(cols) if c in ALIASES}
            if "school_year" not in mapping.values() and "subject" not in mapping.values():
                continue
            for raw in it:
                if not raw or all(c is None for c in raw):
                    continue
                rec = {k: "" for k in CANON}
                for i, name in mapping.items():
                    if i < len(raw) and raw[i] is not None:
                        rec[name] = str(raw[i]).strip()
                if not rec["school_year"]:
                    rec["school_year"] = sheet.strip()
                rec["grade"] = _grade(rec["grade"])
                rec["subject"] = rec["subject"].upper().replace("MATHEMATICS", "MATH")
                if not any(rec[k] for k in ("pct_level_1", "pct_level_2",
                                            "pct_level_3", "pct_level_4")):
                    continue
                yield rec
    finally:
        wb.close()


KEY = ("school_year", "org_no", "bldg_no", "student_group",
       "grade", "subject", "population")


def _rowkey(rec):
    return "\x1f".join(rec[k] for k in KEY)


def main():
    files = sorted(FULL_DIR.glob("*.xlsx"))
    if not files:
        print(f"no workbooks in {FULL_DIR} - run scripts/ks/assessment.py first")
        return
    DEST.parent.mkdir(parents=True, exist_ok=True)
    tmp = DEST.with_suffix(".tmp.csv")

    # pass 1: read every sheet, tagging each row with the workbook it came from
    read, per_file = 0, []
    with open(tmp, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["_src"] + CANON, extrasaction="ignore")
        w.writeheader()
        for idx, f in enumerate(tqdm.tqdm(files, desc="reading")):
            n = 0
            try:
                for rec in read_workbook(f):
                    w.writerow({"_src": idx, **rec})
                    n += 1
            except Exception as e:
                tqdm.tqdm.write(f"  {f.name}: {str(e)[:70]}")
            per_file.append((f.name, n))
            read += n
            tqdm.tqdm.write(f"  {f.name:46s} {n:>9,} rows")

    # pass 2: for each natural key, which workbook published it last
    winner = {}
    with open(tmp, newline="", encoding="utf-8") as fh:
        for rec in tqdm.tqdm(csv.DictReader(fh), desc="deduping", unit=" rows"):
            winner[_rowkey(rec)] = rec["_src"]

    # pass 3: keep only that copy
    kept, seen = 0, set()
    with open(tmp, newline="", encoding="utf-8") as fin, \
         open(DEST, "w", newline="", encoding="utf-8") as fout:
        w = csv.DictWriter(fout, fieldnames=CANON, extrasaction="ignore")
        w.writeheader()
        for rec in csv.DictReader(fin):
            k = _rowkey(rec)
            if winner[k] != rec["_src"] or k in seen:
                continue
            seen.add(k)
            w.writerow(rec)
            kept += 1
    tmp.unlink()

    print(f"\n{read:,} rows read -> {kept:,} after dedup "
          f"({read - kept:,} duplicated across the two-year windows)")
    print(f"{DEST} ({DEST.stat().st_size/1e6:.0f} MB)")
    empty = [n for n, c in per_file if c == 0]
    if empty:
        print("no data (KSDE placeholders): " + ", ".join(empty))


if __name__ == "__main__":
    main()
